from io import BytesIO
from datetime import date, timedelta
from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from ..models.schemas import SchedulingParams, AlgorithmConfig
from ..algorithm.scheduler import MultiProductScheduler


TEMPLATE_PATH = r"d:\Desktop\智能排产\智能排产POC\导出模板.xlsx"

WEEKDAYS = ["星期一", "星期二", "星期三", "星期四", "星期五", "星期六", "星期日"]

HEADER_ROWS = 3
DATA_ROWS = 5
COL_START = 10  # J


def _copy_style(src, dst):
    dst.font = copy(src.font)
    dst.fill = copy(src.fill)
    dst.border = copy(src.border)
    dst.alignment = copy(src.alignment)
    if src.number_format:
        dst.number_format = src.number_format


def _clear_range(ws, min_row, max_row, min_col, max_col):
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.value = None


class SchedulingService:
    @staticmethod
    def calculate(params: SchedulingParams, config: AlgorithmConfig, holidays: list = None) -> dict:
        holiday_list = holidays or []
        scheduler = MultiProductScheduler(
            products=params.products,
            start_date=params.start_date,
            end_date=params.end_date,
            holidays=holiday_list,
            max_time_seconds=config.max_time_seconds,
            rest_day_weight=config.rest_day_weight,
            max_consecutive_work_days=config.max_consecutive_work_days,
        )
        return scheduler.run()

    @staticmethod
    def export_excel(result: dict, plan_info: list = None) -> BytesIO:
        if plan_info and hasattr(plan_info[0], 'model_dump'):
            plan_info = [p.model_dump() for p in plan_info]
        daily_results = result.get("daily_results", [])
        num_days = len(daily_results)
        num_products = result.get("num_products", 1)

        wb = load_workbook(TEMPLATE_PATH)
        ws = wb.active

        pink_fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")

        last_date_col = COL_START + num_days - 1

        for mc in list(ws.merged_cells.ranges):
            if mc.min_row >= HEADER_ROWS + 1:
                ws.unmerge_cells(str(mc))
            elif mc.min_col >= COL_START:
                ws.unmerge_cells(str(mc))

        total_data_rows = num_products * DATA_ROWS
        existing_data_rows = ws.max_row - HEADER_ROWS

        need_insert = total_data_rows - existing_data_rows
        if need_insert > 0:
            ws.insert_rows(HEADER_ROWS + existing_data_rows + 1, amount=need_insert)

        def _ref_cell(row_offset, col):
            return ws.cell(row=HEADER_ROWS + 1 + row_offset, column=col)

        def _set_cell(row_offset, col, value):
            return ws.cell(row=HEADER_ROWS + 1 + row_offset, column=col, value=value)

        style_template = {}
        for ro in range(DATA_ROWS):
            for c in range(1, ws.max_column + 1):
                style_template[(ro, c)] = _ref_cell(ro, c)

        _clear_range(ws, HEADER_ROWS + 1, HEADER_ROWS + total_data_rows, 1, ws.max_column)

        for pidx in range(num_products):
            base_ro = pidx * DATA_ROWS
            pi = (plan_info or [{}])[pidx] if isinstance(plan_info, list) else (plan_info or {}).get(f"product_{pidx + 1}", {})

            for ro in range(DATA_ROWS):
                for c in range(1, COL_START):
                    src = style_template.get((ro, c))
                    dst = _set_cell(base_ro + ro, c, None)
                    if src:
                        _copy_style(src, dst)

            for c in range(COL_START, last_date_col + 1):
                for ro in range(DATA_ROWS):
                    src = style_template.get((ro, COL_START))
                    dst = _set_cell(base_ro + ro, c, None)
                    if src:
                        _copy_style(src, dst)

            net_demand = pi.get("total_delivery", 0) + pi.get("safety_stock", 0) - pi.get("initial_inventory", 0)
            _set_cell(base_ro, 1, pidx + 1)
            _set_cell(base_ro, 2, pi.get("code", ""))
            _set_cell(base_ro, 3, pi.get("name", ""))
            _set_cell(base_ro, 4, pi.get("initial_inventory", 0))
            _set_cell(base_ro, 5, pi.get("safety_stock", 0))
            _set_cell(base_ro, 6, net_demand)
            _set_cell(base_ro, 7, pi.get("rated_output", 0))

            _set_cell(base_ro, 8, "A")
            _set_cell(base_ro + 1, 8, "B")
            _set_cell(base_ro + 2, 8, "计划生产量")
            _set_cell(base_ro + 3, 8, "计划交货量")
            _set_cell(base_ro + 4, 8, "库存结存量")

            for c in [1, 2, 3, 4, 5, 6, 7]:
                if DATA_ROWS > 1:
                    ws.merge_cells(
                        start_row=HEADER_ROWS + 1 + base_ro, start_column=c,
                        end_row=HEADER_ROWS + 1 + base_ro + DATA_ROWS - 1, end_column=c
                    )

            sum_shift1 = 0
            sum_shift2 = 0
            sum_output = 0

            for i in range(num_days):
                col = COL_START + i
                dr = daily_results[i]
                p = dr["products"][pidx]

                prod1 = p.get("prod1", 0)
                prod2 = p.get("prod2", 0)
                output = p.get("daily_output", 0)
                delivery = p.get("daily_delivery", 0)
                inv = p.get("closing_inventory", 0)

                _set_cell(base_ro, col, prod1 if prod1 > 0 else None)
                _set_cell(base_ro + 1, col, prod2 if prod2 > 0 else None)
                _set_cell(base_ro + 2, col, output)
                _set_cell(base_ro + 3, col, delivery)
                _set_cell(base_ro + 4, col, inv)

                sum_shift1 += prod1
                sum_shift2 += prod2
                sum_output += output

            _set_cell(base_ro, 9, sum_shift1 if sum_shift1 > 0 else None)
            _set_cell(base_ro + 1, 9, sum_shift2 if sum_shift2 > 0 else None)
            _set_cell(base_ro + 2, 9, sum_output)
            _set_cell(base_ro + 3, 9, pi.get("total_delivery", 0))

        start_date_val = date.fromisoformat(daily_results[0]["date"])
        ws.cell(row=3, column=4, value=start_date_val).number_format = "m/d"
        num_weeks = (num_days + 6) // 7

        for w in range(num_weeks):
            w_start = COL_START + w * 7
            w_end = min(COL_START + (w + 1) * 7 - 1, last_date_col)
            if w_start <= w_end:
                if w_start < w_end:
                    start_ref = f"{get_column_letter(w_start)}1"
                    end_ref = f"{get_column_letter(w_end)}1"
                    ws.merge_cells(f"{start_ref}:{end_ref}")
                week_num = (start_date_val + timedelta(weeks=w)).isocalendar()[1]
                cell = ws.cell(row=1, column=w_start, value=f"{week_num}周")
                ref = style_template.get((0, COL_START))
                if ref:
                    cell.font = copy(ref.font)
                    cell.alignment = copy(ref.alignment)

        for i in range(num_days):
            col = COL_START + i
            d = start_date_val + timedelta(days=i)
            ws.cell(row=2, column=col, value=d).number_format = "m/d"
            ws.cell(row=3, column=col, value=WEEKDAYS[d.weekday()])

        for i in range(num_days):
            col = COL_START + i
            dr = daily_results[i]
            if dr.get("is_rest", False):
                ws.cell(row=2, column=col).fill = pink_fill
                ws.cell(row=3, column=col).fill = pink_fill

        total_rows = HEADER_ROWS + total_data_rows
        cell_font = Font(name="宋体", size=9)
        cell_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        for r in range(1, total_rows + 1):
            for c in range(COL_START, last_date_col + 1):
                cell = ws.cell(row=r, column=c)
                cell.font = copy(cell_font)
                cell.border = copy(thin_border)
                cell.alignment = copy(cell_alignment)

        for w in range(num_weeks):
            w_start = COL_START + w * 7
            cell = ws.cell(row=1, column=w_start)
            cell.font = copy(cell_font)
            cell.border = copy(thin_border)
            cell.alignment = copy(cell_alignment)

        max_col = COL_START + num_days
        ws.freeze_panes = ws.cell(row=HEADER_ROWS + 1, column=COL_START)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
